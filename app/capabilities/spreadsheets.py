from __future__ import annotations
from pathlib import Path
class SpreadsheetCapability:
    def inspect(self,path:str)->dict:
        import openpyxl
        wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
        return {"sheets":wb.sheetnames,"rows":{s:wb[s].max_row for s in wb.sheetnames},"columns":{s:wb[s].max_column for s in wb.sheetnames}}
